import qrcode
import pytz
import logging
import base64
import openpyxl
import xmlrpc.client
import unicodedata
import requests
import os
import io
import time
import traceback
from django.contrib import messages
from azure.storage.blob import BlobServiceClient
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from .forms import CtrlCapacitacionesForm, RegistrationForm
from .models import CtrlCapacitaciones, EventImage
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse, FileResponse
from django.urls import reverse
from io import BytesIO
from django.db import connections
from dotenv import load_dotenv
from datetime import datetime
from urllib.parse import quote, unquote, urlparse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.views.decorators.csrf import csrf_exempt
from .utils import registrar_log_interno
from app_learning.services.odoo_conection import fetch_departametos_from_odoo
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from PyPDF2 import PdfReader
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader

logger = logging.getLogger(__name__)

# Cargar variables de entorno
load_dotenv()
database = os.getenv("DATABASE")
user = os.getenv("ODOO_USER")
password = os.getenv("PASSWORD")
host = os.getenv("HOST")
apphost = os.getenv('APP_HOST')

@settings.AUTH.login_required
def index(request, *, context):
    user = context['user']
    return HttpResponse(f"Hello, {user.get('name')}.")

def get_employee_names(request):
    ids = request.GET.getlist('ids[]', [])
    if ids:
        try:
            common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
            uid = common.authenticate(database, user, password, {})
            models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

            # Buscar empleados por identificación
            employees = models.execute_kw(database, uid, password,
                                          'hr.employee', 'search_read',
                                          [[['identification_id', 'in', ids]]],
                                          {'fields': ['id', 'name', 'identification_id']})

            results = [{'id': emp['identification_id'], 'name': emp['name']} for emp in employees]
            return JsonResponse({'results': results})
        except Exception as e:
            logger.error('Failed to fetch employee names from Odoo', exc_info=True)
            return JsonResponse({'error': 'Error fetching employee names'}, status=500)
    else:
        return JsonResponse({'results': []})

def upload_to_azure_blob(file, filename):
    logger.info('Intentando subir archivo a Azure Blob Storage...')
    try:
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("AZURE_CONTAINER_NAME")

        if not connection_string or not container_name:
            logger.error("Azure Storage no configurado correctamente (connection_string/container)")
            return None

        # Sanitizar el nombre del archivo para evitar problemas con caracteres especiales
        filename = ''.join(c for c in filename if c.isalnum() or c in '._- ')
        
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=filename)

        # Subir el archivo
        content_settings = None
        if hasattr(file, 'content_type') and file.content_type:
            from azure.storage.blob import ContentSettings
            content_settings = ContentSettings(content_type=file.content_type)
            
        blob_client.upload_blob(file, overwrite=True, content_settings=content_settings)
        logger.info("Archivo subido a Azure Blob Storage")

        return blob_client.url
    except Exception as e:
        logger.error("Error subiendo archivo a Azure Blob Storage", exc_info=True)
        return None

def delete_blob_from_azure(blob_url):
    try:
        # Obtener la cadena de conexión desde las variables de entorno
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        if not connection_string:
            logger.error("Azure Storage connection string no encontrado")
            return False
            
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        
        # Extraer el nombre del contenedor y del blob desde la URL de manera más robusta
        parsed_url = urlparse(blob_url)
        
        # Manejar diferentes formatos de URL de Azure
        if 'blob.core.windows.net' in parsed_url.netloc:
            # Formato estándar: https://account.blob.core.windows.net/container/blob
            path_parts = parsed_url.path.lstrip('/').split('/', 1)
        else:
            # Podría ser una URL personalizada o CDN
            logger.error("Formato de URL no reconocido para Azure Blob URL", extra={'blob_url': blob_url})
            return False
        
        if len(path_parts) < 2:
            logger.error("URL de blob malformada", extra={'blob_url': blob_url})
            return False
            
        container_name = path_parts[0]
        blob_name = unquote(path_parts[1])


        
        # Verificar si el contenedor existe
        try:
            container_client = blob_service_client.get_container_client(container_name)
            container_client.get_container_properties()
        except Exception as e:
            logger.error("Contenedor de Azure no existe", extra={'container_name': container_name})
            return False
        
        # Obtener el cliente del blob
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
        
        # Verificar si el blob existe antes de intentar eliminarlo
        try:
            blob_client.get_blob_properties()
        except Exception as e:
            logger.error("Blob no existe en Azure", extra={'blob_name': blob_name})
            return False
        
        # Eliminar el blob
        blob_client.delete_blob()
        
        logger.info("Blob eliminado exitosamente")
        return True
    except Exception as e:
        logger.error("Error eliminando blob de Azure", exc_info=True)
        return False
        
# Vista para eliminar imagenes
def delete_image(request, image_id):
    try:
        image = get_object_or_404(EventImage, id=image_id)
        capacitacion_id = image.capacitacion.id

        if request.method == 'POST':
            # Guardar la URL antes de eliminar la imagen
            image_url = image.image_url
            
            # Primero eliminar la instancia de la imagen de la base de datos
            image.delete()
            
            # Luego intentar eliminar el blob de Azure
            success = delete_blob_from_azure(image_url)
            if success:
                messages.success(request, "La imagen ha sido eliminada exitosamente.")
            else:
                # Aún si falla la eliminación del blob, la imagen ya se eliminó de la base de datos
                messages.warning(request, "La imagen se eliminó de la base de datos, pero hubo un problema al eliminar el archivo de Azure.")
        else:
            messages.error(request, "Solicitud inválida.")
    except Exception as e:
        messages.error(request, f"Hubo un error al eliminar la imagen")
        logger.error("Error en delete_image", exc_info=True)

    return redirect('view_assistants', id=capacitacion_id)

# Nueva vista para mostrar una imagen específica en tamaño completo
def view_image(request, image_id):
    image = get_object_or_404(EventImage, id=image_id)
    capacitacion = image.capacitacion
    
    context = {
        'image': image,
        'capacitacion': capacitacion
    }
    
    return render(request, 'view_image.html', context)

# Conversión a UTC asegurando que el objeto sea datetime
def convert_to_utc(dt, timezone_str):
    if isinstance(dt, datetime):
        local = pytz.timezone(timezone_str)
        local_dt = local.localize(dt, is_dst=None)
        utc_dt = local_dt.astimezone(pytz.utc)
        return utc_dt
    else:
        raise ValueError("El objeto proporcionado no es de tipo datetime.datetime")

# Función para obtener el ID del departamento
def get_department_id(department_name):
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

        department = models.execute_kw(database, uid, password,
            'hr.department', 'search_read',
            [[['name', '=', department_name]]],
            {'fields': ['id'], 'limit': 1})
        
        if department:
            return department[0]['id']
        else:
            return None
    
    except Exception as e:
        logger.error('Failed to fetch department ID from Odoo', exc_info=True)
        return None

# Función para obtener el ID del empleado por # de documento
def get_employee_id_by_name(name):
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

        employees = models.execute_kw(database, uid, password,
            'hr.employee', 'search_read',
            [[['name', '=', name]]],
            {'fields': ['id'], 'limit': 1})
        
        if employees:
            return employees[0]['id']
        else:
            return None
        
    except Exception as e:
        logger.error('Failed to fetch employee ID from Odoo', exc_info=True)
        return None

# Función para actualizar registro existente en Odoo con verificación
def update_record_in_odoo(record_id, update_data, capacitacion_id, employee_id):
    """
    Actualiza un registro existente en Odoo y verifica que la actualización fue exitosa.
    Retorna: (success, employee_name, error_message)
    """
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        
        if not uid:
            return False, None, "No se pudo autenticar con Odoo"
        
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')
        
        # Actualizar el registro
        result = models.execute_kw(database, uid, password, 
                                  'x_capacitacion_emplead', 'write', 
                                  [[record_id], update_data])
        
        if not result:
            logger.error(f"La actualización no retornó éxito. Record ID: {record_id}")
            return False, None, "La actualización no fue exitosa"
        
        logger.info(f"Registro actualizado en Odoo. ID: {record_id}")
        
        # Verificar que la actualización fue exitosa leyendo el registro
        time.sleep(0.5)  # Pequeña pausa para asegurar que Odoo procesó la actualización
        
        updated_records = models.execute_kw(database, uid, password,
                                           'x_capacitacion_emplead', 'search_read',
                                           [[
                                               ['id', '=', record_id],
                                               ['x_studio_id_capacitacion', '=', capacitacion_id],
                                               ['x_studio_many2one_field_iphhw', '=', employee_id]
                                           ]],
                                           {'fields': ['id', 'x_studio_asisti'], 'limit': 1})
        
        if not updated_records or len(updated_records) == 0:
            logger.error(f"CRÍTICO: Registro no encontrado después de actualización. ID: {record_id}")
            return False, None, "No se pudo verificar la actualización del registro"
        
        # Verificar que la asistencia se actualizó correctamente
        if 'x_studio_asisti' in update_data:
            if updated_records[0]['x_studio_asisti'] != update_data['x_studio_asisti']:
                logger.error(f"CRÍTICO: La asistencia no se actualizó correctamente. Esperado: {update_data['x_studio_asisti']}, Obtenido: {updated_records[0]['x_studio_asisti']}")
                return False, None, "La actualización no se aplicó correctamente"
        
        # Obtener el nombre del empleado
        employee_data = models.execute_kw(database, uid, password,
                                         'hr.employee', 'search_read',
                                         [[['id', '=', employee_id]]],
                                         {'fields': ['identification_id'], 'limit': 1})
        
        if not employee_data or 'identification_id' not in employee_data[0]:
            logger.warning(f"No se pudo obtener el nombre del empleado. ID: {employee_id}")
            employee_name = None
        else:
            employee_name = employee_data[0]['identification_id']
        
        logger.info(f"Registro actualizado y verificado exitosamente. Record ID: {record_id}")
        return True, employee_name, None
        
    except Exception as e:
        logger.error(f'Error al actualizar registro en Odoo. ID: {record_id}', exc_info=True)
        return False, None, f"Error al actualizar el registro: {str(e)}"

# Función para verificar que un registro existe en Odoo
def verify_record_in_odoo(record_id, capacitacion_id, employee_id):
    """
    Verifica que un registro realmente existe en Odoo después de crearlo.
    Retorna True si existe, False en caso contrario.
    """
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')
        
        # Buscar el registro por ID y verificar que existe
        records = models.execute_kw(database, uid, password,
                                   'x_capacitacion_emplead', 'search_read',
                                   [[
                                       ['id', '=', record_id],
                                       ['x_studio_id_capacitacion', '=', capacitacion_id],
                                       ['x_studio_many2one_field_iphhw', '=', employee_id]
                                   ]],
                                   {'fields': ['id', 'x_studio_asisti'], 'limit': 1})
        
        if records and len(records) > 0:
            logger.info(f"Registro verificado exitosamente en Odoo. ID: {record_id}")
            return True
        else:
            logger.error(f"Registro no encontrado en Odoo después de creación. ID esperado: {record_id}")
            return False
            
    except Exception as e:
        logger.error(f'Error al verificar registro en Odoo. ID: {record_id}', exc_info=True)
        return False

# Función para enviar datos a Odoo con reintentos y verificación robusta
def send_to_odoo(data, max_retries=3):
    """
    Envía datos a Odoo con reintentos automáticos y verificación post-creación.
    Retorna: (record_id, employee_name, url_reunion, error_message)
    Si hay error, record_id será None y error_message contendrá el mensaje.
    """
    logger.debug("Intentando enviar datos a Odoo", extra={'document_id': data.get('document_id'), 'capacitacion_id': data.get('capacitacion_id')})
    
    last_error = None
    
    for attempt in range(1, max_retries + 1):
        try:
            logger.debug(f"Intento {attempt} de {max_retries} para enviar datos a Odoo")
            
            common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
            uid = common.authenticate(database, user, password, {})
            
            if not uid:
                raise ValueError("No se pudo autenticar con Odoo")
            
            models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

            department_id = get_department_id(data['department'])
            if not department_id:
                raise ValueError(f"Department '{data['department']}' not found in Odoo")

            # Verificar si 'employee_id' está en 'data'; si no, buscarlo
            if 'employee_id' in data and data['employee_id']:
                employee_id = data['employee_id']
            else:
                # Buscar el empleado por 'document_id' si 'employee_id' no está disponible
                employee_data = models.execute_kw(database, uid, password,
                                                  'hr.employee', 'search_read',
                                                  [[['name', '=', data['document_id']]]],
                                                  {'fields': ['id', 'name'], 'limit': 1})
                logger.debug("Respuesta de busqueda de empleado recibida")

                if not employee_data:
                    raise ValueError(f"Empleado con documento '{data['document_id']}' no encontrado en Odoo")

                employee_id = employee_data[0]['id']

            # Convertir fechas a cadenas en UTC
            date_str = data['date'].strftime('%Y-%m-%d')
            start_time_str = data['start_time'].strftime('%H:%M:%S')
            end_time_str = data['end_time'].strftime('%H:%M:%S')
            
            # Preparar los datos para el registro en Odoo
            odoo_data = {
                'x_studio_tema': data['topic'],
                'x_studio_many2one_field_iphhw': employee_id,
                'x_studio_fecha_sesin': date_str,
                'x_studio_hora_inicial': start_time_str,
                'x_studio_hora_final': end_time_str,
                'x_studio_many2one_field_ftouu': department_id,
                'x_studio_estado': 'ACTIVA',
                'x_studio_modalidad': data.get('mode', ''),
                'x_studio_ubicacin': data.get('location', ''),
                'x_studio_url': data.get('url_reunion', ''),
                'x_studio_asisti': 'Si',
                'x_studio_tipo': data.get('tipo', ''),
                'x_studio_fecha_hora_registro': data['registro_datetime'],
                'x_studio_ip_del_registro': data.get('ip_address', ''),
                'x_studio_user_agent': data.get('user_agent', ''),
                'x_studio_longitud': data.get('longitude', ''),
                'x_studio_latitud': data.get('latitude', ''),
                'x_studio_moderador': data.get('moderator', ''),
                'x_studio_responsable': data.get('in_charge', ''),
                'x_studio_id_capacitacion': data['capacitacion_id']
            }

            # Verificar que no haya valores None en los datos antes de enviar
            for key, value in odoo_data.items():
                if value is None:
                    raise ValueError(f"El campo {key} tiene un valor None, lo que no es permitido en Odoo.")
            
            # Crear el registro en Odoo
            record_id = models.execute_kw(database, uid, password,
                                          'x_capacitacion_emplead', 'create', [odoo_data])
            
            if not record_id:
                raise ValueError("Odoo no retornó un ID de registro válido después de la creación")

            logger.info(f"Registro creado en Odoo. ID: {record_id}", extra={
                'record_id': record_id,
                'employee_id': employee_id,
                'capacitacion_id': data['capacitacion_id']
            })

            # VERIFICACIÓN CRÍTICA: Verificar que el registro realmente existe en Odoo
            time.sleep(0.5)  # Pequeña pausa para asegurar que Odoo procesó la creación
            
            record_verified = verify_record_in_odoo(record_id, data['capacitacion_id'], employee_id)
            
            if not record_verified:
                logger.error(f"CRÍTICO: Registro creado pero no verificado en Odoo. ID: {record_id}")
                raise ValueError("El registro fue creado pero no se pudo verificar en Odoo. Por favor, intente nuevamente.")

            # Obtener el nombre del empleado desde el campo `identification_id`
            employee_data = models.execute_kw(database, uid, password,
                                              'hr.employee', 'search_read',
                                              [[['id', '=', employee_id]]],
                                              {'fields': ['identification_id'], 'limit': 1})
            
            if not employee_data or 'identification_id' not in employee_data[0]:
                logger.warning(f"No se pudo obtener el nombre del empleado. ID: {employee_id}")
                employee_name = data.get('document_id', 'Usuario')
            else:
                employee_name = employee_data[0]['identification_id']

            logger.info(f"Registro completado exitosamente. Record ID: {record_id}, Employee: {employee_name}")
            return record_id, employee_name, data.get('url_reunion', ''), None

        except (xmlrpc.client.Fault, ConnectionError, TimeoutError) as e:
            # Errores transitorios que pueden reintentarse
            last_error = str(e)
            logger.warning(f"Error transitorio al enviar datos a Odoo (intento {attempt}/{max_retries}): {last_error}")
            if attempt < max_retries:
                time.sleep(1 * attempt)  # Backoff exponencial
                continue
            else:
                logger.error(f'Failed to send data to Odoo after {max_retries} attempts', exc_info=True)
                return None, None, None, f"Error de conexión con el sistema. Por favor, intente nuevamente en unos momentos."
                
        except ValueError as e:
            # Errores de validación que no deben reintentarse
            last_error = str(e)
            logger.error(f'Error de validación al enviar datos a Odoo: {last_error}', exc_info=True)
            return None, None, None, f"Error al procesar los datos: {last_error}"
            
        except Exception as e:
            # Otros errores
            last_error = str(e)
            logger.error(f'Error inesperado al enviar datos a Odoo (intento {attempt}/{max_retries}): {last_error}', exc_info=True)
            if attempt < max_retries:
                time.sleep(1 * attempt)
                continue
            else:
                return None, None, None, f"Error inesperado al registrar la asistencia. Por favor, contacte al administrador."
    
    # Si llegamos aquí, todos los intentos fallaron
    return None, None, None, f"No se pudo registrar la asistencia después de {max_retries} intentos. Por favor, intente nuevamente."

# Función para crear QR de Capacitación
@csrf_exempt
@settings.AUTH.login_required()
def create_capacitacion(request, *, context):


    if request.method == 'POST':
        request.POST = request.POST.copy()
        request.POST['estado'] = 'ACTIVA'
        form = CtrlCapacitacionesForm(request.POST, request.FILES)
        if form.is_valid():
            capacitacion = form.save(commit=False)
            capacitacion.estado = 'ACTIVA'
            capacitacion.verificacion_identidad = form.cleaned_data['verificacion_identidad']
            capacitacion.save()

            
            if capacitacion.tipo in ['Reunión', 'Capacitación']:
                # Procesar archivo PDF si se envió
                if 'archivo_pdf' in request.FILES:
                    logger.info("Archivo PDF detectado para carga en Azure")
                    pdf_file = request.FILES['archivo_pdf']

                    if pdf_file.content_type == 'application/pdf':
                        filename = f"pdf_evento_{capacitacion.id}_{pdf_file.name}"
                        logger.info("Subiendo archivo PDF a Azure", extra={'filename': filename})

                        pdf_url = upload_to_azure_blob(pdf_file, filename)

                        if pdf_url:
                            logger.info("PDF subido correctamente a Azure")
                            capacitacion.pdf_url = pdf_url
                            capacitacion.save()
                        else:
                            logger.error("Fallo la carga de PDF a Azure: url vacia")
                            messages.error(request, "Error al subir el archivo PDF a Azure.")
                    else:
                        logger.warning("Tipo de archivo inválido para PDF", extra={'content_type': pdf_file.content_type})
                        messages.error(request, "El archivo debe ser un PDF.")
                else:
                    logger.debug("No se adjunto archivo PDF en la solicitud")

            employee_ids = request.POST.get('employee_names', '').split(',')
            logger.debug("Ids de empleados seleccionados recibidos en POST") 
            
                        
            if employee_ids:
                logger.info("Enviando asistentes preseleccionados a Odoo")
                send_assistants_to_odoo(capacitacion.id, employee_ids)
                     
            qr_url = f"{apphost}/learn/register/?id={capacitacion.id}"

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_url)
            qr.make(fit=True)
            img = qr.make_image(fill='black', back_color='white')

            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
            capacitacion.qr_base64 = img_base64
            capacitacion.save()
            
            userdata = context['user']
            username = userdata.get('name')
            email = userdata.get('preferred_username')
            
            #Log Data
            username = email
            observacion = (f"Creación de la capacitación: {capacitacion.id}")
            id= capacitacion.id
            tipo = "Creación"
            
            #Create Log
            registrar_log_interno(username, observacion, tipo, id)
            
            logger.info("Capacitacion creada", extra={'capacitacion_id': capacitacion.id, 'estado': capacitacion.estado})
            return HttpResponseRedirect(reverse('details_view', args=[capacitacion.id]))
        else:
            logger.warning("Formulario de creacion de capacitacion invalido", extra={'errors': form.errors.as_json()})
    else:
        form = CtrlCapacitacionesForm()
    return render(request, 'crear_capacitacion.html', {'form': form})

#Duplicar una capacitación
@settings.AUTH.login_required()
def duplicate_event(request, id, *, context):
    original = get_object_or_404(CtrlCapacitaciones, id=id)
    
    if request.method == 'POST':
        logger.info("Solicitud POST para duplicar capacitacion", extra={'original_id': original.id})
        # Forzar estado ACTIVA (el campo puede venir deshabilitado y no enviarse)
        post_data = request.POST.copy()
        post_data['estado'] = 'ACTIVA'
        form = CtrlCapacitacionesForm(post_data, request.FILES)
        if form.is_valid():
            try:
                nueva = form.save(commit=False)
                # Forzar estado ACTIVA en duplicacion
                nueva.estado = 'ACTIVA'
                # Asegurar verificacion_identidad desde el form (evitar None)
                nueva.verificacion_identidad = form.cleaned_data.get('verificacion_identidad', getattr(nueva, 'verificacion_identidad', 'NO'))
                nueva.save()

                # Generar QR para la nueva capacitacion
                qr_url = f"{apphost}/learn/register/?id={nueva.id}"
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(qr_url)
                qr.make(fit=True)
                img = qr.make_image(fill='black', back_color='white')
                buffer = BytesIO()
                img.save(buffer, format='PNG')
                buffer.seek(0)
                nueva.qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                nueva.save()

                # Log de auditoria interno
                userdata = context['user']
                email = userdata.get('preferred_username')
                registrar_log_interno(email, f"Duplicacion de capacitacion: {original.id} -> {nueva.id}", "Creación", nueva.id)

                messages.success(request, "La capacitación fue duplicada exitosamente.")
                logger.info("Capacitacion duplicada exitosamente", extra={'original_id': original.id, 'nueva_id': nueva.id})
                return HttpResponseRedirect(reverse('details_view', args=[nueva.id]))
            except Exception:
                logger.error("Error creando la capacitacion duplicada", exc_info=True, extra={'original_id': original.id})
                messages.error(request, "No fue posible crear la copia de la capacitación.")
        else:
            logger.warning("Formulario invalido al duplicar capacitacion", extra={'original_id': original.id, 'errors': form.errors.as_json()})
            messages.error(request, "Datos inválidos al duplicar la capacitación. Verifique los campos requeridos.")
    else:
        initial_data = {
            'fecha': original.fecha,
            'tipo': original.tipo,
            'privacidad': original.privacidad,
            'tema': original.tema,
            'responsable': original.responsable,
            'moderador': original.moderador,
            'hora_inicial': original.hora_inicial,
            'hora_final': original.hora_final,
            'total_invitados': original.total_invitados,
            'area_encargada': original.area_encargada,
            'modalidad': original.modalidad,
            'objetivo': original.objetivo,
            'estado': 'ACTIVA',  # Forzar como activa
            'verificacion_identidad': original.verificacion_identidad,
            'url_reunion': original.url_reunion,
            'ubicacion': original.ubicacion,
            'temas': original.temas,
        }

        form = CtrlCapacitacionesForm(initial=initial_data)

    return render(request, 'crear_capacitacion.html', {'form': form})

# Función para mostrar lista de Capacitaciones
def list_capacitaciones(request):
    capacitaciones = CtrlCapacitaciones.objects.all()
    return render(request, 'list_capacitaciones.html', {'capacitaciones': capacitaciones})


def password_validation(identificacion: str, password_md5: str) -> bool:
    """
    Valida que la identificación y contraseña (md5) coincidan
    con un registro en la tabla 'intranet_empleados_usuarios' de la base secundaria.

    Args:
        identificacion (str): Número de identificación (campo login).
        password_md5 (str): Contraseña en MD5 a validar.

    Returns:
        bool: True si la validación es correcta, False en caso contrario.
    """
    with connections['auth_db'].cursor() as cursor:
        query = """
            SELECT COUNT(*) FROM sitio_web.INTRANET_EMPLEADOS_USUARIOS
            WHERE login = %s AND contrasena = %s
        """
        cursor.execute(query, [identificacion, password_md5])
        result = cursor.fetchone()

    return result[0] > 0

# Función para registrar asistencia y actualizar registro en Odoo
def registration_view(request, id=None):
    if(id):
        capacitacion_id = id
        capacitacion = get_object_or_404(CtrlCapacitaciones, id=capacitacion_id)
    else:    
        capacitacion_id = request.GET.get('id')
        capacitacion = get_object_or_404(CtrlCapacitaciones, id=capacitacion_id)

    # Formatea la fecha correctamente
    date_str = capacitacion.fecha.strftime('%Y-%m-%d')

    initial_data = {
        'topic': capacitacion.tema,
        'objective': capacitacion.objetivo,
        'department': capacitacion.area_encargada,
        'moderator': capacitacion.moderador,
        'tipo': capacitacion.tipo,
        'date': date_str,
        'start_time': capacitacion.hora_inicial,  # Formato de 24 horas
        'end_time': capacitacion.hora_final,      # Formato de 24 horas
        'mode': capacitacion.modalidad,
        'location': capacitacion.ubicacion,
        'url_reunion': capacitacion.url_reunion,
        'in_charge': capacitacion.responsable,
        'privacidad': capacitacion.privacidad,
        'document_id': ''  # Este campo se llenará por el usuario
    }


    form = RegistrationForm(request.POST or None, initial=initial_data)
    
    is_active = capacitacion.estado == 'ACTIVA'

    error_message = None  # Inicializa la variable error_message

    if request.method == 'POST' and is_active:
        
        if form.is_valid():
            document_id = form.cleaned_data['document_id']
            if 'hashed_password' in form.cleaned_data:
                

                #llamar la función para verificar la contraseña
                # Llamar a la función password_validation
                #password_validation(document_id, hashed_password)
                                # Verificar si se requiere contraseña o no
                if capacitacion.verificacion_identidad == "NO" or (capacitacion.verificacion_identidad == "SI" and password_validation(document_id, form.cleaned_data['hashed_password'])):
                    
                    logger.info('Validacion de identidad correcta')
                    try:
                        # Verifica la privacidad del evento
                        if capacitacion.privacidad == "CERRADA": #priv
                            # Verifica si el documento existe en Odoo
                            employee_id = get_employee_id_by_name(document_id)

                            if not employee_id:
                                error_message = f"No se encontró un empleado con el documento {document_id}. Por favor, verifique los datos."
                            else:
                                common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
                                uid = common.authenticate(database, user, password, {})
                                models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

                                # Buscar si ya existe un registro del asistente en la capacitación en Odoo
                                existing_records = models.execute_kw(database, uid, password,
                                    'x_capacitacion_emplead', 'search_read',
                                    [[
                                        ['x_studio_id_capacitacion', '=', capacitacion_id],
                                        ['x_studio_many2one_field_iphhw', '=', employee_id]
                                    ]],
                                    {'fields': ['id', 'x_studio_asisti']})
                                
                                timezone = pytz.timezone('America/Bogota')
                                registro_datetime = datetime.now(timezone).strftime('%Y-%m-%d %H:%M:%S')
                                user_agent = request.META.get('HTTP_USER_AGENT', '')
                                ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR'))
                                latitude = request.POST.get('latitude')
                                longitude = request.POST.get('longitude')
                                
                                if ip_address:
                                    ip_address = ip_address.split(',')[0]
                                else:
                                    ip_address = request.META.get('REMOTE_ADDR')

                                if existing_records:
                                    # Si el registro ya existe, verificar si la asistencia está en "No"
                                    record_id = existing_records[0]['id']
                                    asistencia_actual = existing_records[0]['x_studio_asisti']

                                    if asistencia_actual == 'No':
                                        update_data = {
                                            'x_studio_asisti': 'Si',
                                            'x_studio_fecha_hora_registro': registro_datetime,
                                            'x_studio_ip_del_registro': ip_address,
                                            'x_studio_user_agent': user_agent,
                                            'x_studio_longitud': longitude or '',
                                            'x_studio_latitud': latitude or '',
                                        }
                                        
                                        # Usar función robusta para actualizar con verificación
                                        success, employee_name, update_error = update_record_in_odoo(
                                            record_id, update_data, capacitacion_id, employee_id
                                        )
                                        
                                        if success and employee_name:
                                            encoded_url = quote(capacitacion.url_reunion or 'without-url', safe='')
                                            return redirect(reverse('success', kwargs={'employee_name': employee_name, 'url_reunion': encoded_url}))
                                        else:
                                            # Si la actualización falla, mostrar página de error
                                            logger.error(f"Error al actualizar registro. ID: {record_id}, Error: {update_error}")
                                            return render(request, 'registration_error.html', {
                                                'error_message': update_error or "No se pudo actualizar el registro de asistencia.",
                                                'responsable': capacitacion.responsable,
                                                'capacitacion_id': capacitacion_id
                                            })  
                                    else:
                                        error_message = f"El usuario con documento {document_id} ya ha registrado su asistencia."

                                else:
                                    # Redirigir al template de alerta si no está inscrito
                                    return render(request, 'no_inscrito.html', {
                                        'responsable': capacitacion.responsable, 
                                        'capacitacion_id': capacitacion_id,
                                        'area_encargada':capacitacion.area_encargada
                                        })

                        else:  # Si la privacidad es 'ABIERTA', continuar con el flujo normal
                            common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
                            uid = common.authenticate(database, user, password, {})
                            models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

                            employee_id = get_employee_id_by_name(document_id)
                            if not employee_id:
                                error_message = f"No se encontró un empleado con el documento {document_id}. Por favor, verifique los datos."
                            else:
                                # Obtener datos adicionales
                                timezone = pytz.timezone('America/Bogota')
                                registro_datetime = datetime.now(timezone).strftime('%Y-%m-%d %H:%M:%S')
                                user_agent = request.META.get('HTTP_USER_AGENT', '')
                                ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR'))
                                latitude = request.POST.get('latitude')
                                longitude = request.POST.get('longitude')

                                if ip_address:
                                    ip_address = ip_address.split(',')[0]
                                else:
                                    ip_address = request.META.get('REMOTE_ADDR')

                                # Verificar si ya existe un registro del asistente en la capacitación en Odoo
                                existing_records = models.execute_kw(database, uid, password,
                                    'x_capacitacion_emplead', 'search_read',
                                    [[
                                        ['x_studio_id_capacitacion', '=', capacitacion_id],
                                        ['x_studio_many2one_field_iphhw', '=', employee_id]
                                    ]],
                                    {'fields': ['id', 'x_studio_asisti']})

                                if existing_records:
                                    # Si el registro ya existe
                                    record_id = existing_records[0]['id']
                                    asistencia_actual = existing_records[0]['x_studio_asisti']

                                    if asistencia_actual == 'No':
                                        update_data = {
                                            'x_studio_asisti': 'Si',
                                            'x_studio_fecha_hora_registro': registro_datetime,
                                            'x_studio_ip_del_registro': ip_address,
                                            'x_studio_user_agent': user_agent,
                                            'x_studio_longitud': longitude or '',
                                            'x_studio_latitud': latitude or '',
                                        }

                                        # Usar función robusta para actualizar con verificación
                                        success, employee_name, update_error = update_record_in_odoo(
                                            record_id, update_data, capacitacion_id, employee_id
                                        )
                                        
                                        if success and employee_name:
                                            encoded_url = quote(capacitacion.url_reunion or 'without-url', safe='')
                                            return redirect(reverse('success', kwargs={'employee_name': employee_name, 'url_reunion': encoded_url}))
                                        else:
                                            # Si la actualización falla, mostrar página de error
                                            logger.error(f"Error al actualizar registro. ID: {record_id}, Error: {update_error}")
                                            return render(request, 'registration_error.html', {
                                                'error_message': update_error or "No se pudo actualizar el registro de asistencia.",
                                                'responsable': capacitacion.responsable,
                                                'capacitacion_id': capacitacion_id
                                            })
                                    else:
                                        error_message = f"El usuario con documento {document_id} ya ha registrado su asistencia."
                                else:
                                    # Si no existe el registro, crear uno nuevo
                                    data = form.cleaned_data
                                    data['registro_datetime'] = registro_datetime
                                    data['ip_address'] = ip_address
                                    data['user_agent'] = user_agent
                                    data['latitude'] = latitude
                                    data['longitude'] = longitude
                                    data['capacitacion_id'] = capacitacion_id
                                    data['employee_id'] = employee_id  # Pasar el ID del empleado

                                    # Usar función robusta con verificación post-creación
                                    record_id, employee_name, url_reunion, error_msg = send_to_odoo(data)
                                    
                                    if record_id and employee_name:
                                        # Verificación exitosa - redirigir a success
                                        encoded_url = quote(url_reunion or 'without-url', safe='')
                                        return redirect(reverse('success', kwargs={'employee_name': employee_name, 'url_reunion': encoded_url}))
                                    else:
                                        # Error al crear registro - mostrar página de error
                                        logger.error(f"Error al crear registro en Odoo. Documento: {document_id}, Error: {error_msg}")
                                        return render(request, 'registration_error.html', {
                                            'error_message': error_msg or "No se pudo registrar la asistencia en el sistema.",
                                            'responsable': capacitacion.responsable,
                                            'capacitacion_id': capacitacion_id
                                        })


                    except Exception as e:
                        logger.error('Error inesperado al registrar la asistencia en Odoo:', exc_info=True, extra={
                            'document_id': document_id,
                            'capacitacion_id': capacitacion_id,
                            'error': str(e)
                        })
                        # Mostrar página de error en lugar de solo un mensaje
                        return render(request, 'registration_error.html', {
                            'error_message': f"Error inesperado al procesar el registro: {str(e)}. Por favor, contacte al administrador.",
                            'responsable': capacitacion.responsable,
                            'capacitacion_id': capacitacion_id
                        })
                else:
                    return render(request, 'registration_form.html', {
                        'form': form,
                        'capacitacion': capacitacion,
                        'is_active': is_active,
                        'error_message': 'Documento o contraseña incorrectos. Por favor, verifique sus datos.'
                    })


        else:
            logger.warning("Formulario de registro invalido", extra={'errors': form.errors.as_json()})

    context = {
        'form': form,
        'is_active': is_active,
        'capacitacion': capacitacion,
        'error_message': error_message
    }

    return render(request, 'registration_form.html', context)

# vista para obtener la configuracion de verificacion de identidad de una capacitacion.
def verificacion_config(request):
    capacitacion_id = request.GET.get('id')
    try:
        capacitacion = CtrlCapacitaciones.objects.get(id=capacitacion_id)
        return JsonResponse({
            'verificacion_identidad': capacitacion.verificacion_identidad if
            hasattr(capacitacion, 'verificacion_identidad') else 'NO'
        })
    except CtrlCapacitaciones.DoesNotExist:
        return JsonResponse({'error': 'capacitacion no encontrada'},
        status=404)
# Vista de Éxito Al Enviar Datos

def success_view(request, employee_name, url_reunion=None):
    decoded_url = unquote(url_reunion) if url_reunion and url_reunion != 'without-url' else None
    context = {
        'employee_name': employee_name,
        'url_reunion': decoded_url  # Usa la URL decodificada o None si no se proporciona
    }
    return render(request, 'success.html', context)

#Vista Detalles de la Capacitación
@settings.AUTH.login_required()
def details_view(request, id, *, context):
    capacitacion = get_object_or_404(CtrlCapacitaciones, id=id)
    
     # Determinar si mostrar ubicación y/o URL de la reunión según la modalidad
    show_url = capacitacion.modalidad == 'VIRTUAL' or capacitacion.modalidad == 'MIXTA'
    show_ubicacion = capacitacion.modalidad == 'PRESENCIAL' or capacitacion.modalidad == 'MIXTA'
    
    fecha_str = capacitacion.fecha.strftime('%Y-%m-%d')
    fecha_fin_str = capacitacion.fecha_fin.strftime('%Y-%m-%d') if capacitacion.fecha_fin else None
    
    if fecha_fin_str and fecha_str != fecha_fin_str:
        date_display = f"{fecha_str} al {fecha_fin_str}"
    else:
        date_display = fecha_str
    
    context = {
        'topic': capacitacion.tema,
        'department': capacitacion.area_encargada,
        'in_charge': capacitacion.responsable,
        'objective': capacitacion.objetivo,
        'moderator': capacitacion.moderador,
        'date': date_display,
        'start_time': capacitacion.hora_inicial.strftime('%H:%M'),
        'end_time': capacitacion.hora_final.strftime('%H:%M'),
        'modalidad': capacitacion.modalidad, 
        'ubicacion': capacitacion.ubicacion if show_ubicacion else None,  # Condicionalmente según la modalidad
        'url_reunion': capacitacion.url_reunion if show_url else None,  # Condicionalmente según la modalidad
        'qr_url': f"{apphost}/learn/register/?id={capacitacion.id}",
        'qr_base64': capacitacion.qr_base64,
        'topics': capacitacion.temas,
    }
    
    return render(request, 'details_view.html', context)

# Vista Home que muestra todas las capacitaciones
@settings.AUTH.login_required()
def home(request, *, context):
    capacitaciones = CtrlCapacitaciones.objects.all().order_by('estado','-fecha', '-hora_inicial')
    for capacitacion in capacitaciones:
        capacitacion.fecha_formateada = capacitacion.fecha.strftime('%Y-%m-%d')
        capacitacion.hora_inicial_formateada = capacitacion.hora_inicial.strftime('%H:%M')
        capacitacion.hora_final_formateada = capacitacion.hora_final.strftime('%H:%M')
    departments = fetch_departametos_from_odoo() or []
    return render(request, 'home.html', {
        'capacitaciones': capacitaciones,
        'departments': departments
    })

#Actualizar Capacitación en Odoo por ID:
def update_odoo_capacitacion (capacitacion):
    try:
        #Conexión Odoo
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')
        
        odoo_capacitaciones_ids = models.execute_kw(database, uid, password,
            'x_capacitacion_emplead', 'search',
            [[['x_studio_id_capacitacion', '=', capacitacion.id]]])
        
        
        if odoo_capacitaciones_ids:
            update_data = {
                'x_studio_tema': capacitacion.tema,
                'x_studio_fecha_sesin': capacitacion.fecha.strftime('%Y-%m-%d'),
                'x_studio_hora_inicial': capacitacion.hora_inicial.strftime('%H:%M:%S'),
                'x_studio_hora_final': capacitacion.hora_final.strftime('%H:%M:%S'),
                'x_studio_estado': capacitacion.estado,
                'x_studio_modalidad': capacitacion.modalidad,
                'x_studio_ubicacin': capacitacion.ubicacion or '',
                'x_studio_url': capacitacion.url_reunion or '',
                'x_studio_moderador': capacitacion.moderador,
                'x_studio_tipo': capacitacion.tipo,
                'x_studio_responsable': capacitacion.responsable,
            }
            
            models.execute_kw(database, uid, password,
                              'x_capacitacion_emplead', 'write',
                              [odoo_capacitaciones_ids, update_data])
            
            logger.info(f"Capacitación con ID {capacitacion.id} actualizada en Odoo para {len(odoo_capacitaciones_ids)} registros.")
            
        else:
            logger.warning(f"No se encontró la capacitación con ID {capacitacion.id} en Odoo")
               
        
    except Exception as e:
        logger.error('Error actualizando datos en Odoo', exc_info=True)
        

# Vista para editar una capacitación existente
@csrf_exempt
@settings.AUTH.login_required()
def edit_capacitacion(request, id, *, context):
    capacitacion = get_object_or_404(CtrlCapacitaciones, id=id)
    userdata = context['user']
    username = userdata.get('name')
    email = userdata.get('preferred_username')
    
    if request.method == 'POST':
        form = CtrlCapacitacionesForm(request.POST, instance=capacitacion)
        if form.is_valid():
            capacitacion = form.save(commit=False)
            capacitacion.user = username
            # Asegurarse de que el campo verificacion_identidad se guarde correctamente
            # capacitacion.verificacion_identidad = form.cleaned_data['verificacion_identidad']
            form.save()
            # Obtener los empleados seleccionados del POST
            employee_names = request.POST.get('employee_names', '').split(',')
            logger.debug("Asistentes seleccionados en edicion recibidos")

            # Llamar a send_assistants_to_odoo después de guardar la capacitación
            if employee_names:
                logger.info("Enviando asistentes a Odoo desde edicion")
                send_assistants_to_odoo(capacitacion.id, employee_names)
            
            #Log Data
            username = email
            id= capacitacion.id
            
            if capacitacion.estado == 'ACTIVA':
                observacion = f"Modificación de la capacitación: {capacitacion.id}"
                tipo = "Actualización"
            else:
                observacion = f"Cierre de la capacitación: {capacitacion.id}"
                tipo = "Cierre"

            
            #Create Log
            registrar_log_interno(username, observacion, tipo, id)
            #Update Odoo
            update_odoo_capacitacion(capacitacion)
            return redirect('home')
    else:
        form = CtrlCapacitacionesForm(instance=capacitacion)
    return render(request, 'crear_capacitacion.html', {'form': form})

# Vista para ver los usuarios que asistieron a una capacitación
@settings.AUTH.login_required()
def view_assistants(request, id, *, context):
    capacitacion = get_object_or_404(CtrlCapacitaciones, id=id)
    error_message = None
    success_message = None
    
    
    if request.method == 'POST' and 'archivo_presentacion' in request.FILES:
        archivo = request.FILES['archivo_presentacion']
        
        #Verficar el tipo de archivo
        if archivo.content_type not in ['application/pdf', 'application/vnd.openxmlformats-officedocument.presentationml.presentation']:
            messages.error(request, "Formato no permitido. Solo se aceptan archivos PDF o PPTX.")
        else:
            filename = f"presentacion_{capacitacion.id}_{archivo.name}"
            file_url = upload_to_azure_blob(archivo, filename)

            if file_url:
                capacitacion.archivo_presentacion = file_url
                capacitacion.save()
                messages.success(request, "Archivo cargado correctamente.")
            else:
                messages.error(request, "Error al subir el archivo.")
    
    if request.method == 'POST' and 'image' in request.FILES:
        image_file = request.FILES['image']
        
        
        #Verificar el tamaño de archivo
        if image_file.size > 3 * 1024 *1024:
            error_message = "La imagen excede los 3MB"
        else:
            filename = f"evidencia_{capacitacion.id}_{image_file.name}"
            
            #Cargar imagen a Azure Blob Storage
            image_url = upload_to_azure_blob(image_file, filename)
            
            if image_url:
                # Guardar la URL de la img en el modelo
                capacitacion.image_url = image_url
                capacitacion.save()
                success_message = "Imagen cargada con éxito"
            else:
                error_message = "Erro al cargar la imagen"
        if error_message:
            messages.error(request, error_message)
        if success_message:
            messages.success(request, success_message)
    
    #Remover Acentos
    def remove_accents(input_str):
        nfkd_form = unicodedata.normalize('NFKD', input_str)
        return "".join([c for c in nfkd_form if not unicodedata.combining(c)])
        
    capacitacion = get_object_or_404(CtrlCapacitaciones, id=id)
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

        assistants = models.execute_kw(database, uid, password,
            'x_capacitacion_emplead', 'search_read',
            [[['x_studio_id_capacitacion', '=',id]]],
            {'fields': ['x_studio_many2one_field_iphhw', 'x_studio_cargo', 'x_studio_nombre_empleado', 'x_studio_departamento_empleado',
                        'x_studio_correo_personal', 'x_studio_correo_corporativo', 'x_studio_asisti']})

        assistant_data_yes = []
        assistant_data_no = []
        
        for assistant in assistants:
            userId = assistant['x_studio_many2one_field_iphhw'][1] if assistant['x_studio_many2one_field_iphhw'] else ''
            jobTitle = assistant.get('x_studio_cargo', '')
            username = assistant.get('x_studio_nombre_empleado','')
            employeeDepartment = assistant.get('x_studio_departamento_empleado','')
            personalEmail = assistant.get('x_studio_correo_personal')
            corporateEmail = assistant.get('x_studio_correo_corporativo') 
            x_studio_asisti = assistant.get('x_studio_asisti','')
            
            
            assistant_data = {
                'userId': userId, 
                'jobTitle': jobTitle, 
                'username':username, 
                'employeeDepartment': employeeDepartment,
                'personalEmail':personalEmail,
                'corporateEmail':corporateEmail
            }

            if x_studio_asisti == 'Si':
                assistant_data_yes.append(assistant_data)
            elif x_studio_asisti == 'No':
                assistant_data_no.append(assistant_data)
        
        #Calcular porcentaje de asistencia:
        total_invitados = capacitacion.total_invitados
        total_asistentes = len(assistant_data_yes)
        tasa_exito = 0
        total_ausentes = total_invitados - total_asistentes
        
        if total_invitados > 0 :
            tasa_exito = (total_asistentes/total_invitados)*100
        
        if request.method == 'POST' and 'images' in request.FILES:
            images = request.FILES.getlist('images')
            total_existing_images = capacitacion.images.count()
            total_images = total_existing_images + len(images)
            
            if total_images > 5:
                messages.error(request, "No puede tener más de 5 imágenes en total")
            else:
                for image_file in images:
                    if image_file.size > 3 * 1024 * 1024:
                        messages.error(request, f"La imagen {image_file.name} excede el tamaño máximo de 3MB.")
                        continue  # Saltar esta imagen y continuar con las demás

                    filename = f"capacitacion_{capacitacion.id}_{image_file.name}"
                    image_url = upload_to_azure_blob(image_file, filename)
                    logger.debug('Imagen cargada a Azure para capacitacion', extra={'capacitacion_id': capacitacion.id})
                    if image_url:
                        EventImage.objects.create(capacitacion=capacitacion, image_url=image_url)
                        messages.success(request, f"La imagen {image_file.name} ha sido cargada exitosamente.")
                    else:
                        messages.error(request, f"No se pudo subir la imagen {image_file.name}.")
                    
        
        # Función para formatear los encabezados
        def format_excel_headers(ws):
            # Color verde #5C9C31 en los encabezados
            fill = PatternFill(start_color="5C9C31", end_color="5C9C31", fill_type="solid")
            # Fuente en negrita
            font = Font(bold=True, color= "FFFFFF")
            # Alineación centrada
            alignment = Alignment(horizontal="center", vertical="center")

            # Aplicar formato a cada celda del encabezado (primera fila)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment

            # Ajustar automáticamente el ancho de las columnas basado en el texto más largo en todas las filas
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)  # Obtener la letra de la columna
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))  # Buscar el texto más largo en todas las filas
                adjusted_width = max_length + 5  # Agregar un poco de espacio adicional
                ws.column_dimensions[column_letter].width = adjusted_width  # Ajustar el ancho de la columna

        # Aplicación en el lugar correcto donde se crea el archivo Excel
        if request.GET.get('download') == 'excel':
            # Crear archivo de Excel en memoria
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = remove_accents(f"Datos Asistentes")

            # Escribir encabezados
            ws.append(["Número de Documento", "Nombre", "Cargo", "Área", "Correo Personal", "Correo Corporativo"])

           

            # Agregar datos de los asistentes con manejo de valores vacíos
            for assistant in assistant_data_yes:
                row = [
                    assistant['userId'].strip() if assistant['userId'] else '',
                    assistant['username'].strip() if assistant['username'] else '',
                    assistant['jobTitle'].strip() if assistant['jobTitle'] else '',
                    assistant['employeeDepartment'].strip() if assistant['employeeDepartment'] else '',
                    assistant['personalEmail'].strip() if assistant['personalEmail'] else '',
                    assistant['corporateEmail'].strip() if assistant['corporateEmail'] else ''
                ]
                ws.append(row)
                
            # Formatear archivo
            format_excel_headers(ws)

            # Guardar el archivo en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Generar respuesta HTTP con el archivo
            sanitized_filename = remove_accents(f"Asistentes_{capacitacion.tema}.xlsx")
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={sanitized_filename}'

            # Cerrar el workbook antes de enviar la respuesta
            wb.close()
            return response

    except Exception as e:
        logger.error('Failed to fetch assistants from Odoo', exc_info=True)
        assistant_data = []
        tasa_exito = 0

    return render(request, 'view_assistants.html', {
        'capacitacion': capacitacion,
        'assistants_yes': assistant_data_yes,
        'assistants_no': assistant_data_no,
        'total_invitados': total_invitados,
        'tasa_exito': round(tasa_exito, 2), # Cifra redondeada
        'total_ausentes': total_ausentes,
        'error_message': error_message,
        'success_message': success_message
    })
    
def generar_pdf(request, id):
    # 1. Obtener la capacitación y los asistentes
    capacitacion = get_object_or_404(CtrlCapacitaciones, id=id)
    asistentes_data = get_asistentes_odoo(capacitacion.id) or []
    
    #Rango de fechas
    
    fecha_str = capacitacion.fecha.strftime('%Y-%m-%d') if capacitacion.fecha else "No disponible"
    fecha_fin_str = capacitacion.fecha_fin.strftime('%Y-%m-%d') if capacitacion.fecha_fin else None
    if capacitacion.fecha_fin and fecha_str != fecha_fin_str:
        date_display = f"{fecha_str} al {fecha_fin_str}"
    else:
        date_display = fecha_str

    # 2. Obtener la ruta correcta del membrete

    membrete_path = os.path.join(settings.MEDIA_ROOT, "membrete", "Membrete.pdf")

    # 3. Crear buffer y documento base
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=110,  # Aumentamos la margen superior para el membrete
        bottomMargin=50
    )

    # 4. Definir estilos
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = 'Helvetica'
    styles['Normal'].fontSize = 12
    styles['Normal'].leading = 14

    bold_style = ParagraphStyle('BoldStyle', parent=styles['Normal'], fontName='Helvetica-Bold')
    title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontName='Helvetica-Bold', alignment=1, fontSize=12, leading=14)
    bullet_style = ParagraphStyle('BulletStyle', parent=styles['Normal'], leading=18)

    def P(text, style=styles['Normal']):
        return Paragraph(text if text else "No disponible", style)

    elements = []

    # 5. Título principal
    elements.append(Paragraph(f"FUNDACIÓN UNIVERSIDAD DE ANTIOQUIA<br/><br/>INFORME - {capacitacion.tema or 'No disponible'}", title_style))
    elements.append(Spacer(1, 12))

    # 6. Objetivo
    elements.append(Paragraph("Objetivo:", bold_style))
    elements.append(Spacer(1, 4))
    elements.append(P(capacitacion.objetivo or "No disponible"))
    elements.append(Spacer(1, 12))

    # 7. Tabla con datos principales
    info_data = [
        [P("<b>Evento:</b>"), P(capacitacion.tema)],
        [P("<b>Responsable:</b>"), P(capacitacion.responsable)],
        [P("<b>Moderador:</b>"), P(capacitacion.moderador)],
        [P("<b>Fecha:</b>"), P(date_display)],
        [P("<b>Hora:</b>"), P(f"{capacitacion.hora_inicial.strftime('%H:%M')} - {capacitacion.hora_final.strftime('%H:%M')}")]
    ]

    if capacitacion.modalidad in ["PRESENCIAL", "MIXTA"] and capacitacion.ubicacion:
        info_data.append([P("<b>Lugar:</b>"), P(capacitacion.ubicacion)])
    if capacitacion.modalidad in ["VIRTUAL", "MIXTA"] and capacitacion.url_reunion:
        info_data.append([P("<b>URL Reunión:</b>"), P(capacitacion.url_reunion)])

    info_table = Table(info_data, colWidths=[120, 420])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    # 8. Temas
    if capacitacion.temas:
        elements.append(Paragraph("Temas:", bold_style))
        elements.append(Spacer(1, 4))
        for tema in capacitacion.temas.split(","):
            elements.append(Paragraph(f"• {tema.strip()}", bullet_style))
        elements.append(Spacer(1, 12))

    # 9. Lista de asistentes
    # 1️⃣ Lista de asistentes (solo si hay datos)
    if asistentes_data:
        elements.append(Paragraph("Lista de Asistentes:", bold_style))
        if capacitacion.verificacion_identidad == "SI":
            elements.append(Paragraph("**Este listado de asistentes tiene validacion de identidad con el usuario y contraseña de la intranet**", styles['Italic']))
        elements.append(Spacer(1, 4))

        asistentes_table_data = [
            [P("<b>#</b>"), P("<b>Identificación</b>"), P("<b>Nombre</b>"), P("<b>Cargo</b>"), P("<b>Área</b>"), P("<b>Fecha registro</b>")]
        ]
        for idx, assistant in enumerate(asistentes_data, start=1):
            asistentes_table_data.append([
                P(str(idx)),
                P(assistant.get('employeeId', 'No disponible')),
                P(assistant.get('username', 'No disponible')),
                P(assistant.get('jobTitle', 'No disponible')),
                P(assistant.get('employeeDepartment') or assistant.get('employeeCompany') or 'no hay'),
                P(assistant.get('registroFechaHora', 'No disponible'))
            ])
        total_width = letter[0] - 80  

        asistentes_table = Table(asistentes_table_data, colWidths=[
            total_width * 0.08,  # numero
            total_width * 0.18,  # identificacion
            total_width * 0.22,  # nombre
            total_width * 0.18,  # cargo
            total_width * 0.18,  # area
            total_width * 0.16   # fecha registro
        ])
        asistentes_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#D4EDDA")),  # Verde claro
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Texto negro
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(asistentes_table)
        elements.append(Spacer(1, 12))

        # 10. Evidencias del Evento
        


    # 10a. Presentación (mostrar “ver aquí”)
    if capacitacion.archivo_presentacion:
        elements.append(Paragraph("Evidencias del Evento:", bold_style))
        elements.append(Spacer(1, 6))
        presentacion_paragraph = Paragraph(
            f'Presentación del evento: <font color="blue"><u><link href="{capacitacion.archivo_presentacion}">ver aquí</link></u></font>',
            styles['Normal']
        )
        elements.append(presentacion_paragraph)
        elements.append(Spacer(1, 12))

    # 10b. Imágenes
    event_images = capacitacion.images.all()
    if event_images.exists():
        elements.append(Paragraph("Imágenes de evidencia:", bold_style))
        elements.append(Spacer(1, 4))
        for image in event_images:
            try:
                resp = requests.get(image.image_url)
                if resp.status_code == 200:
                    img_data = io.BytesIO(resp.content)
                    img_obj = Image(img_data)
                    img_obj._restrictSize(500, 300)
                    elements.append(img_obj)
                    elements.append(Spacer(1, 12))
                else:
                    elements.append(P(f"Error al cargar imagen: {image.image_url}"))
            except Exception:
                elements.append(P(f"Error al cargar imagen: {image.image_url}"))
            elements.append(Spacer(1, 6))

    # 11. Función para agregar membrete
    def add_membrete(canvas, doc):
        try:
            with open(membrete_path, "rb") as f:
                reader = PdfReader(f)
                page = reader.pages[0]
                membrete_img = ImageReader(io.BytesIO(page.images[0].data))

            canvas.drawImage(membrete_img, 0, 0, width=612, height=792, preserveAspectRatio=True)
        except Exception as e:
            logger.error("Error cargando membrete para PDF", exc_info=True)

    # 12. Generar PDF con membrete en todas las páginas
    doc.build(elements, onFirstPage=add_membrete, onLaterPages=add_membrete)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f"{capacitacion.tema or 'Reporte'}.pdf")# Buscar Empleados en Odoo

#Buscar Empleados
def search_employees(request):
    query = request.GET.get('q', '')
    search_type = request.GET.get('search_type', 'id') #Por defecto busca por ID
    
    if not query:
        return JsonResponse({'results': []})
    
    common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
    uid = common.authenticate(database, user, password, {})
    models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

    # Construir el dominio de busqueda segun el tipo de busqueda
    if search_type == 'id':
        # Buscar solo por numero de identificacion 
        domain = [('identification_id', 'ilike', query)]
    elif search_type == 'name':
        # buscar solo por nombre
        domain = [('name', 'ilike', query)]
    else: # 'both'
        # Buscar por nommbre como por identificacion
        domain =['|', ('identification_id', 'ilike', query), ('name', 'ilike', query)]
    # Esto ejecuta la busqueda en Odoo y obtiene los resultados en forma de diccionarios
    employee_ids = models.execute_kw(database, uid, password,
        'hr.employee', 'search_read',
        [domain],
        {'fields': ['name', 'identification_id'], 'limit': 10})

    return JsonResponse({'results': employee_ids})

#Enviar Asistentes Obligatorios a Odoo
def send_assistants_to_odoo(capacitacion_id, employee_ids):
    try:
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')
        
        capacitacion = CtrlCapacitaciones.objects.get(id=capacitacion_id)
        
        # Obtener el ID del departamento (area encargada) en Odoo
        department_data = models.execute_kw(database, uid, password,
                                            'hr.department', 'search_read',
                                            [[['name', '=', capacitacion.area_encargada]]],
                                            {'fields': ['id'], 'limit': 1})
        
        if not department_data:
            raise ValueError(f"No se encontró el departamento '{capacitacion.area_encargada}' en Odoo.")
        
        department_id = department_data[0]['id']
        
        
        
       # Loop sobre cada empleado seleccionado
        logger.info('Enviando asistentes a Odoo')
        for name in employee_ids:
            employee_data = models.execute_kw(database, uid, password,
                                              'hr.employee', 'search_read',
                                              [[['name', '=', name]]],
                                              {'fields': ['id'], 'limit': 1})
            if employee_data:
                employee_id = employee_data[0]['id']  # Obtener el ID del empleado en Odoo
                odoo_data = {
                    'x_studio_tema': capacitacion.tema,
                    'x_studio_many2one_field_iphhw': employee_id,  # ID del empleado en Odoo
                    'x_studio_fecha_sesin': capacitacion.fecha.strftime('%Y-%m-%d'),
                    'x_studio_hora_inicial': capacitacion.hora_inicial.strftime('%H:%M:%S'),
                    'x_studio_hora_final': capacitacion.hora_final.strftime('%H:%M:%S'),                                                                                                                       
                    'x_studio_estado':capacitacion.estado,
                    'x_studio_many2one_field_ftouu': department_id,
                    'x_studio_asisti': 'No',  # Marcamos asistencia en 'No'
                    'x_studio_id_capacitacion': capacitacion_id,  # Relacionar con la capacitación
                    'x_studio_responsable': capacitacion.responsable,
                    'x_studio_moderador': capacitacion.moderador,
                    'x_studio_tipo': capacitacion.tipo,
                    'x_studio_modalidad': capacitacion.modalidad,
                    'x_studio_ubicacin': capacitacion.ubicacion or '',
                    'x_studio_url': capacitacion.url_reunion or ''
                }
                
                # Crear el registro de asistente en Odoo
                models.execute_kw(database, uid, password, 'x_capacitacion_emplead', 'create', [odoo_data])

        logger.info(f"Asistentes enviados a Odoo para la capacitación {capacitacion_id}")
    except Exception as e:
        logger.error('Failed to send assistants to Odoo', exc_info=True)

#Obtener Asistentes registrados a un evento en Odoo     
def get_asistentes_odoo(capacitacion_id):
    try:
        # Conexión a Odoo
        common = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/common')
        uid = common.authenticate(database, user, password, {})
        models = xmlrpc.client.ServerProxy(f'{host}/xmlrpc/2/object')

        # Buscar asistentes
        records = models.execute_kw(
            database, uid, password,
            'x_capacitacion_emplead', 'search_read',
            [[['x_studio_id_capacitacion', '=', capacitacion_id]]],
            {
                'fields': [
                    'x_studio_nombre_empleado', 
                    'x_studio_cargo', 
                    'x_studio_departamento_empleado',
                    'x_studio_capacitacion_compania_empleado',
                    'x_studio_many2one_field_iphhw',
                    'x_studio_fecha_hora_registro'
                    ],
                'limit': 999
            }
        )

        # Ajustar data
        asistentes_data = []
        for record in records:
            asistentes_data.append({
                'username': record.get('x_studio_nombre_empleado', ''),
                'jobTitle': record.get('x_studio_cargo', ''),
                'employeeDepartment': record.get('x_studio_departamento_empleado'),
                'employeeCompany': record.get('x_studio_capacitacion_compania_empleado'),
                'employeeId': record.get('x_studio_many2one_field_iphhw', [])[1] if record.get('x_studio_many2one_field_iphhw') else '',
                'registroFechaHora': record.get('x_studio_fecha_hora_registro') or ''
                
            })
        return asistentes_data

    except Exception as e:
        logger.error("Error obteniendo asistentes de Odoo", exc_info=True)
        return []